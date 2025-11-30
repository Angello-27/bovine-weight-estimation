"""
Excel Generator - Core Layer
Generador de reportes en formato Excel con estilos profesionales y personalizados

Single Responsibility: Generar archivos Excel con formatos avanzados, colores y diseño
"""

import io
from datetime import datetime
from typing import Any

from openpyxl import Workbook  # type: ignore[import-untyped, import-not-found, import]
from openpyxl.styles import (  # type: ignore[import-untyped, import-not-found, import]
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)


class ExcelGenerator:
    """
    Generador de reportes Excel con estilos profesionales.

    Single Responsibility: Crear archivos Excel con estilos avanzados, colores y formato.
    """

    # Paleta de colores - Tema Light App Móvil (Hacienda Gamelera)
    COLOR_HEADER = "255946"  # Verde marca principal (headers)
    COLOR_PRIMARY = "255946"  # Verde marca principal
    COLOR_PRIMARY_LIGHT = "49A760"  # Verde claro
    COLOR_PRIMARY_DARK = "1F4E3D"  # Verde oscuro
    COLOR_ACCENT = "EFB443"  # Dorado accent
    COLOR_SUCCESS = "49A760"  # Verde claro (success)
    COLOR_ERROR = "EF4444"  # Rojo (error)
    COLOR_WARNING = "EFB443"  # Dorado (warning)
    COLOR_LIGHT = "FAFAFA"  # Surface (gris muy claro)
    COLOR_DARK = "212121"  # Texto principal (gris oscuro)
    COLOR_BACKGROUND = "FFFFFF"  # Blanco (background)
    COLOR_TEXT_SECONDARY = "757575"  # Gris texto secundario

    @staticmethod
    def _apply_header_style(cell: Any) -> None:
        """Aplica estilo a celdas de encabezado."""
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(
            start_color=ExcelGenerator.COLOR_HEADER,
            end_color=ExcelGenerator.COLOR_HEADER,
            fill_type="solid",
        )
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000"),
        )

    @staticmethod
    def _apply_data_style(cell: Any, is_alternate: bool = False) -> None:
        """Aplica estilo a celdas de datos."""
        cell.font = Font(size=10)
        if is_alternate:
            cell.fill = PatternFill(
                start_color=ExcelGenerator.COLOR_LIGHT,
                end_color=ExcelGenerator.COLOR_LIGHT,
                fill_type="solid",
            )
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC"),
        )

    @staticmethod
    def _apply_title_style(cell: Any) -> None:
        """Aplica estilo a título."""
        cell.font = Font(bold=True, size=16, color=ExcelGenerator.COLOR_PRIMARY)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    @staticmethod
    def generate_traceability_report(data: dict) -> bytes:
        """Genera reporte Excel de trazabilidad con diseño profesional."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Trazabilidad"

        # Título
        ws.merge_cells("A1:B1")
        title_cell = ws["A1"]
        ExcelGenerator._apply_title_style(title_cell)
        title_cell.value = "📋 REPORTE DE TRAZABILIDAD"
        ws.row_dimensions[1].height = 30

        # Información del animal
        row = 3
        animal = data["animal"]

        # Encabezado de sección
        ws.merge_cells(f"A{row}:B{row}")
        section_cell = ws[f"A{row}"]
        section_cell.value = "INFORMACIÓN DEL ANIMAL"
        section_cell.font = Font(bold=True, size=12, color=ExcelGenerator.COLOR_HEADER)
        row += 1

        # Datos del animal
        animal_data = [
            ("Caravana:", animal.ear_tag),
            ("Raza:", animal.breed),
            ("Género:", animal.gender),
            ("Fecha de Nacimiento:", animal.birth_date.strftime("%Y-%m-%d")),
            ("Edad (meses):", animal.calculate_age_months()),
            ("Estado:", animal.status.upper()),
        ]

        if animal.name:
            animal_data.insert(1, ("Nombre:", animal.name))

        for label, value in animal_data:
            label_cell = ws[f"A{row}"]
            value_cell = ws[f"B{row}"]
            label_cell.value = label
            value_cell.value = value
            label_cell.font = Font(bold=True)
            ExcelGenerator._apply_data_style(label_cell)
            ExcelGenerator._apply_data_style(value_cell, is_alternate=(row % 2 == 0))
            row += 1

        row += 1

        # Resumen de pesos
        ws.merge_cells(f"A{row}:B{row}")
        section_cell = ws[f"A{row}"]
        section_cell.value = "RESUMEN DE PESOS"
        section_cell.font = Font(bold=True, size=12, color=ExcelGenerator.COLOR_HEADER)
        row += 1

        summary = data["summary"]
        summary_data = [
            ("Total de Estimaciones:", summary["total_weight_estimations"]),
            (
                "Peso Actual (kg):",
                (
                    f"{summary['current_weight']:.2f}"
                    if summary["current_weight"]
                    else "N/A"
                ),
            ),
            (
                "Primer Peso (kg):",
                f"{summary['first_weight']:.2f}" if summary["first_weight"] else "N/A",
            ),
        ]

        for label, value in summary_data:
            label_cell = ws[f"A{row}"]
            value_cell = ws[f"B{row}"]
            label_cell.value = label
            value_cell.value = value
            label_cell.font = Font(bold=True)
            ExcelGenerator._apply_data_style(label_cell)
            ExcelGenerator._apply_data_style(value_cell, is_alternate=(row % 2 == 0))
            row += 1

        # Ajustar ancho de columnas
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 30

        # Footer
        ws.merge_cells(f"A{row+2}:B{row+2}")
        footer_cell = ws[f"A{row+2}"]
        footer_cell.value = (
            f"Generado el {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')}"
        )
        footer_cell.font = Font(
            size=8, italic=True, color=ExcelGenerator.COLOR_TEXT_SECONDARY
        )
        footer_cell.alignment = Alignment(horizontal="center")

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()

    @staticmethod
    def generate_inventory_report(data: dict) -> bytes:
        """Genera reporte Excel de inventario con diseño profesional."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventario"

        # Título
        ws.merge_cells("A1:F1")
        title_cell = ws["A1"]
        ExcelGenerator._apply_title_style(title_cell)
        title_cell.value = "📊 REPORTE DE INVENTARIO"
        ws.row_dimensions[1].height = 30

        row = 3

        # Estadísticas por raza
        stats = data["statistics"]
        if stats["by_breed"]:
            # Encabezado de tabla
            headers = ["RAZA", "CANTIDAD", "PORCENTAJE"]
            total = data["total_animals"]

            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col)
                cell.value = header
                ExcelGenerator._apply_header_style(cell)

            row += 1

            # Datos
            for breed, count in sorted(
                stats["by_breed"].items(), key=lambda x: x[1], reverse=True
            ):
                percentage = (count / total * 100) if total > 0 else 0
                breed_cell = ws.cell(row=row, column=1)
                count_cell = ws.cell(row=row, column=2)
                pct_cell = ws.cell(row=row, column=3)

                breed_cell.value = breed
                count_cell.value = count
                pct_cell.value = f"{percentage:.1f}%"

                ExcelGenerator._apply_data_style(
                    breed_cell, is_alternate=(row % 2 == 0)
                )
                ExcelGenerator._apply_data_style(
                    count_cell, is_alternate=(row % 2 == 0)
                )
                count_cell.alignment = Alignment(horizontal="center")
                ExcelGenerator._apply_data_style(pct_cell, is_alternate=(row % 2 == 0))
                pct_cell.alignment = Alignment(horizontal="center")
                row += 1

            row += 1

        # Total
        total_cell = ws.cell(row=row, column=1)
        total_cell.value = f"TOTAL DE ANIMALES: {data['total_animals']}"
        total_cell.font = Font(bold=True, size=12, color=ExcelGenerator.COLOR_SUCCESS)
        row += 2

        # Ajustar ancho de columnas
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 15

        # Footer
        ws.merge_cells(f"A{row}:C{row}")
        footer_cell = ws[f"A{row}"]
        footer_cell.value = (
            f"Generado el {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')}"
        )
        footer_cell.font = Font(
            size=8, italic=True, color=ExcelGenerator.COLOR_TEXT_SECONDARY
        )
        footer_cell.alignment = Alignment(horizontal="center")

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()

    @staticmethod
    def generate_movements_report(data: dict) -> bytes:
        """Genera reporte Excel de movimientos con diseño profesional."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Movimientos"

        # Título
        ws.merge_cells("A1:B1")
        title_cell = ws["A1"]
        ExcelGenerator._apply_title_style(title_cell)
        title_cell.value = "🔄 REPORTE DE MOVIMIENTOS"
        ws.row_dimensions[1].height = 30

        row = 3

        # Resumen
        summary = data["summary"]
        summary_data = [
            ("Total de Movimientos:", summary["total_movements"]),
            ("Vendidos:", summary["total_sold"]),
            ("Fallecidos:", summary["total_deceased"]),
        ]

        for label, value in summary_data:
            label_cell = ws[f"A{row}"]
            value_cell = ws[f"B{row}"]
            label_cell.value = label
            value_cell.value = value
            label_cell.font = Font(bold=True)
            ExcelGenerator._apply_data_style(label_cell)
            ExcelGenerator._apply_data_style(value_cell, is_alternate=(row % 2 == 0))
            row += 1

        # Ajustar ancho de columnas
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 20

        # Footer
        row += 2
        ws.merge_cells(f"A{row}:B{row}")
        footer_cell = ws[f"A{row}"]
        footer_cell.value = (
            f"Generado el {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')}"
        )
        footer_cell.font = Font(
            size=8, italic=True, color=ExcelGenerator.COLOR_TEXT_SECONDARY
        )
        footer_cell.alignment = Alignment(horizontal="center")

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()

    @staticmethod
    def generate_growth_report(data: dict) -> bytes:
        """Genera reporte Excel de crecimiento con diseño profesional."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Crecimiento"

        # Título
        ws.merge_cells("A1:B1")
        title_cell = ws["A1"]
        ExcelGenerator._apply_title_style(title_cell)
        title_cell.value = "📈 REPORTE DE CRECIMIENTO"
        ws.row_dimensions[1].height = 30

        row = 3

        if data["type"] == "individual":
            metrics = data["growth_metrics"]
            growth_data = [
                (
                    "GDP (Ganancia Diaria Promedio):",
                    f"{metrics['gdp']:.2f} kg/día" if metrics["gdp"] else "N/A",
                ),
                ("Total de Mediciones:", metrics["total_measurements"]),
                (
                    "Peso Actual (kg):",
                    (
                        f"{metrics['current_weight']:.2f}"
                        if metrics["current_weight"]
                        else "N/A"
                    ),
                ),
                (
                    "Ganancia Total (kg):",
                    (
                        f"{metrics['weight_gain']:.2f}"
                        if metrics["weight_gain"]
                        else "N/A"
                    ),
                ),
            ]

            for label, value in growth_data:
                label_cell = ws[f"A{row}"]
                value_cell = ws[f"B{row}"]
                label_cell.value = label
                value_cell.value = value
                label_cell.font = Font(bold=True)
                ExcelGenerator._apply_data_style(label_cell)
                ExcelGenerator._apply_data_style(
                    value_cell, is_alternate=(row % 2 == 0)
                )
                row += 1

        # Ajustar ancho de columnas
        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 20

        # Footer
        row += 2
        ws.merge_cells(f"A{row}:B{row}")
        footer_cell = ws[f"A{row}"]
        footer_cell.value = (
            f"Generado el {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')}"
        )
        footer_cell.font = Font(
            size=8, italic=True, color=ExcelGenerator.COLOR_TEXT_SECONDARY
        )
        footer_cell.alignment = Alignment(horizontal="center")

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()
